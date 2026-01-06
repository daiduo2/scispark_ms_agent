import os
from common.utils.tool import get_related_keyword, remove_number_prefix, extract_hypothesis, search_releated_paper, extract_technical_entities, extract_message, paper_compression, review_mechanism, extract_message_review, extract_message_review_moa
from common.utils.arxiv_api import search_paper
from common.utils.llm_api import call_with_deepseek, call_with_qwenmax
from common.core.prompt import fact_extraction_prompt, hypothesis_generate_prompt, hypotheses_prompt, initial_idea_prompt, technical_optimizatio_prompt, MoA_based_optimization_prompt, human_ai_collaboration_prompt
from openpyxl import load_workbook, Workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
from common.core.moa import moa_idea_iteration, moa_table
from common.core.config import OUTPUT_PATH
from common.utils.scholar_download import sanitize_folder_name

def process_paper(paper):
    if 'abstract' not in paper:
        return None
    user_prompt = f"""Now,please following these rules to extract the factual information from following paper:\ntitile:{paper['title']}\nabstract:\n{paper['abstract']}\n"""
    result = call_with_deepseek(system_prompt=fact_extraction_prompt(), question=user_prompt)
    return {"topic": paper['topic'], "title": paper['title'], "abstract": paper['abstract'], "result": result}

def Fact_Information_Extraction(Keyword, SearchPaperNum, user_id, task):
    """基于主题检索论文并抽取事实信息，生成 Excel 文件
    
    参数:
    - Keyword: 主题名称
    - SearchPaperNum: 检索论文数量
    - user_id: 用户标识
    - task: 任务对象（用于目录归档）
    
    返回:
    - (file_path, Keywords): 事实信息 Excel 路径与相关关键词列表
    """
    Keywords = get_related_keyword(Keyword=Keyword)
    Keywords.append(Keyword)
    papers = search_paper(Keywords=Keywords, Limit=SearchPaperNum)
    task_id = getattr(task, 'request', {}).get('id', 'default_task_id') if task else 'default_task_id'
    topic_dir = sanitize_folder_name(Keyword.replace(" ", "_"))
    file_path_prefix = fr"{OUTPUT_PATH}/{user_id}/{task_id}/{topic_dir}"
    file_path = fr"{file_path_prefix}/fact_information_{Keyword}.xlsx"
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    start_row = ws.max_row + 1
    results = []
    with ThreadPoolExecutor(max_workers=5) as executor:
        future_to_paper = {executor.submit(process_paper, paper): paper for paper in papers}
        for future in as_completed(future_to_paper):
            res = future.result()
            if res:
                results.append(res)
    for res in results:
        ws.cell(row=start_row, column=1, value=res["topic"]) 
        ws.cell(row=start_row, column=2, value=res["title"]) 
        ws.cell(row=start_row, column=3, value=res["abstract"]) 
        ws.cell(row=start_row, column=4, value=res["result"]) 
        start_row += 1
    wb.save(file_path)
    return file_path, Keywords

def Hypothesis_Generate(Keyword, Fact_File_Path, Keywords, shuffle=False, random_num=5, user_id="", task=None):
    """基于事实信息生成多条假设并保存输入与结果
    
    参数:
    - Keyword: 主题名称
    - Fact_File_Path: 事实信息 Excel 路径
    - Keywords: 相关关键词列表
    - shuffle: 是否打乱事实信息
    - random_num: 每个关键词使用的事实数量上限
    - user_id: 用户标识
    - task: 任务对象（用于目录归档）
    
    返回:
    - str: 假设结果 Markdown 文件路径
    """
    wb = load_workbook(filename=Fact_File_Path)
    ws = wb.active
    fact_information = []
    for keyword in Keywords:
        index = 0
        for paper in ws.iter_rows(min_row=2, values_only=False):
            if paper[0].value == keyword and index < random_num:
                temp = remove_number_prefix(paper[3].value)
                fact_information += temp.split('\n')
                index += 1
    if shuffle:
        import random
        random.shuffle(fact_information)
    Known_Information = ""
    index = 0
    for information in fact_information:
        if information != "" and information != "\n":
            Known_Information += f"{index+1}. {information}\n"
            index += 1
    user_prompt = hypothesis_generate_prompt(Keyword=Keyword, Known_Information=Known_Information)
    result = call_with_deepseek(system_prompt="You are a research expert.", question=user_prompt, temperature=1.5)
    task_id = getattr(task, 'request', {}).get('id', 'default_task_id') if task else 'default_task_id'
    topic_dir = sanitize_folder_name(Keyword.replace(" ", "_"))
    file_path_prefix = fr"{OUTPUT_PATH}/{user_id}/{task_id}/{topic_dir}/Hypotheses"
    os.makedirs(file_path_prefix, exist_ok=True)
    with open(fr"{file_path_prefix}/fromfact_{Keyword}_input.md", 'w', encoding='utf-8') as f:
        f.write(user_prompt)
    with open(fr"{file_path_prefix}/fromfact_{Keyword}_result.md", 'w', encoding='utf-8') as f:
        f.write(result)
    return fr"{file_path_prefix}/fromfact_{Keyword}_result.md"

def Initial_Idea(Keyword, SearchPaperNum=5, Compression=True, user_id="", task=None):
    """生成初始方案草案并保存输入与结果
    
    参数:
    - Keyword: 主题名称
    - SearchPaperNum: 检索论文数量
    - Compression: 是否对论文进行内容压缩
    - user_id: 用户标识
    - task: 任务对象（用于目录归档）
    
    返回:
    - str: 初始方案结果 Markdown 文件路径
    """
    fact_file_path, Keywords = Fact_Information_Extraction(Keyword=Keyword, SearchPaperNum=SearchPaperNum, user_id=user_id, task=task)
    hypo_file_path = Hypothesis_Generate(Keyword=Keyword, Fact_File_Path=fact_file_path, Keywords=Keywords, user_id=user_id, task=task)
    hypotheses = extract_hypothesis(file=hypo_file_path, split_section="Hypothesis")
    hypotheses_index, hypotheses_result = hypotheses_prompt(Hypotheses=hypotheses)
    keynum, relatedPaper, keyword_str = search_releated_paper(topic=Keyword, max_paper_num=SearchPaperNum, compression=Compression, user_id=user_id, task=task)
    title_abstract_prompt = ""
    if Compression:
        for index, paper in enumerate(relatedPaper):
            title_abstract_prompt += f"\n# The {index + 1} related paper\n## title\n{paper['title']}\n## abstract\n{paper['abstract']}\n## content\n{paper['compression_result']}\n"
    else:
        for index, paper in enumerate(relatedPaper):
            title_abstract_prompt += f"\n# The {index + 1} related paper\n## title\n{paper['title']}\n## abstract\n{paper['abstract']}\n"
    paper_count = len(relatedPaper)
    user_prompt = initial_idea_prompt(hypotheses_prompt=hypotheses_result, title_abstract_prompt=title_abstract_prompt, keyword_str=keyword_str, hypotheses_index=hypotheses_index, index=paper_count, keynum=keynum)
    task_id = getattr(task, 'request', {}).get('id', 'default_task_id') if task else 'default_task_id'
    topic_dir = sanitize_folder_name(Keyword.replace(" ", "_"))
    file_path_prefix = fr"{OUTPUT_PATH}/{user_id}/{task_id}/{topic_dir}/Idea"
    os.makedirs(file_path_prefix, exist_ok=True)
    with open(fr"{file_path_prefix}/{Keyword}_initial_input.md", 'w', encoding='utf-8') as f:
        f.write(user_prompt)
    initial_idea_result = call_with_qwenmax(question=user_prompt)
    with open(fr"{file_path_prefix}/{Keyword}_initial_result.md", 'w', encoding='utf-8') as f:
        f.write(initial_idea_result)
    return fr"{file_path_prefix}/{Keyword}_initial_result.md"

def Technical_Optimization(Keyword, Initial_Idea_Result_File, Compression=True, user_id="", task=None):
    """执行技术优化阶段，检索相关论文并生成优化草案
    
    参数:
    - Keyword: 主题名称
    - Initial_Idea_Result_File: 初始方案结果文件路径
    - Compression: 是否对论文进行内容压缩
    - user_id: 用户标识
    - task: 任务对象（用于目录归档）
    
    返回:
    - str: 技术优化结果 Markdown 文件路径
    """
    task_id = getattr(task, 'request', {}).get('id', 'default_task_id') if task else 'default_task_id'
    topic_dir = sanitize_folder_name(Keyword.replace(" ", "_"))
    file_path_prefix = fr"{OUTPUT_PATH}/{user_id}/{task_id}/{topic_dir}/Idea"
    os.makedirs(file_path_prefix, exist_ok=True)
    technical_keywords, Initial_Idea_Result = extract_technical_entities(Initial_Idea_Result_File, split_section="Paper Abstract")
    text, target_paper_title = extract_message(Initial_Idea_Result_File, split_section="Paper Title")
    readysearch_key = []
    for index, keyword in enumerate(technical_keywords):
        if index > 3:
            break
        readysearch_key.append(keyword['entity'])
    Papers = search_paper(Keywords=readysearch_key, Limit=2)
    title_abstract_prompt = ""
    relatedPaper = []
    for paper in Papers:
        if Compression:
            try:
                compression_result = paper_compression(doi=paper["doi"], title=paper["title"], topic=Keyword, user_id=user_id, task=task)
            except Exception:
                compression_result = "None"
        else:
            compression_result = "None"
        try:
            relatedPaper.append({"title": paper["title"], "abstract": paper["abstract"], "compression_result": compression_result})
        except:
            pass
    if Compression:
        for index, paper in enumerate(relatedPaper):
            title_abstract_prompt += f"\n# The {index + 1} related paper\n## title\n{paper['title']}\n## abstract\n{paper['abstract']}\n## content\n{paper['compression_result']}\n"
    else:
        for index, paper in enumerate(relatedPaper):
            title_abstract_prompt += f"\n# The {index + 1} related paper\n## title\n{paper['title']}\n## abstract\n{paper['abstract']}\n"
    user_prompt = technical_optimizatio_prompt(title=target_paper_title, title_abstract_prompt=title_abstract_prompt, Initial_Idea_Result=Initial_Idea_Result)
    with open(fr"{file_path_prefix}/{Keyword}_technical_optimization_input.md", 'w', encoding='utf-8') as f:
        f.write(user_prompt)
    idea_iteration_result = call_with_qwenmax(question=user_prompt)
    with open(fr"{file_path_prefix}/{Keyword}_technical_optimization_result.md", 'w', encoding='utf-8') as f:
        f.write(idea_iteration_result)
    return fr"{file_path_prefix}/{Keyword}_technical_optimization_result.md"

def MoA_Based_Optimization(Keyword, Technical_Optimization_Result_File, Compression=True, user_id="", task=None):
    """执行 MoA 基于评审的优化阶段，生成多代理评审及迭代产物
    
    参数:
    - Keyword: 主题名称
    - Technical_Optimization_Result_File: 技术优化结果文件路径
    - Compression: 是否对论文进行内容压缩
    - user_id: 用户标识
    - task: 任务对象（用于目录归档）
    
    返回:
    - str: MoA 结果 Markdown 文件路径
    """
    task_id = getattr(task, 'request', {}).get('id', 'default_task_id') if task else 'default_task_id'
    topic_dir = sanitize_folder_name(Keyword.replace(" ", "_"))
    file_path_prefix = fr"{OUTPUT_PATH}/{user_id}/{task_id}/{topic_dir}/Idea"
    file_path_review_prefix = fr"{OUTPUT_PATH}/{user_id}/{task_id}/{topic_dir}/Review"
    os.makedirs(file_path_prefix, exist_ok=True)
    draft, target_paper_title = extract_message(file=Technical_Optimization_Result_File, split_section="Paper Title")
    information = review_mechanism(topic=topic_dir, draft=draft, user_id=user_id, task=task)
    text, next_optimization = extract_message_review(file=fr"{file_path_review_prefix}/{Keyword}_review.md", split_section="Next Steps for Optimization")
    target_next_optimization = ""
    for index, opt in enumerate(next_optimization.values()):
        target_next_optimization += f"\n{index+1}.{opt}"
    readysearch_key = []
    optimization_keywords = ""
    Papers = search_paper(Keywords=readysearch_key, Limit=2)
    title_abstract_prompt = ""
    relatedPaper = []
    for paper in Papers:
        if Compression:
            try:
                compression_result = paper_compression(doi=paper["doi"], title=paper["title"], topic=Keyword, user_id=user_id, task=task)
            except Exception:
                compression_result = "None"
        else:
            compression_result = "None"
        try:
            relatedPaper.append({"title": paper["title"], "abstract": paper["abstract"], "compression_result": compression_result})
        except:
            pass
    if Compression:
        for index, paper in enumerate(relatedPaper):
            title_abstract_prompt += f"\n# The {index + 1} related paper\n## title\n{paper['title']}\n## abstract\n{paper['abstract']}\n## content\n{paper['compression_result']}\n"
    else:
        for index, paper in enumerate(relatedPaper):
            title_abstract_prompt += f"\n# The {index + 1} related paper\n## title\n{paper['title']}\n## abstract\n{paper['abstract']}\n"
    user_prompt = MoA_based_optimization_prompt(target_next_optimization=target_next_optimization, optimization_keywords=optimization_keywords, title_abstract_prompt=title_abstract_prompt, draft=draft)
    with open(fr"{file_path_prefix}/{Keyword}_moa_input.md", 'w', encoding='utf-8') as f:
        f.write(user_prompt)
    idea_iteration_result = moa_idea_iteration(topic=Keyword, user_prompt=user_prompt, user_id=user_id, task=task)
    with open(fr"{file_path_prefix}/{Keyword}_moa_result.md", 'w', encoding='utf-8') as f:
        f.write(idea_iteration_result)
    return fr"{file_path_prefix}/{Keyword}_moa_result.md"

def Human_AI_Collaboration(Keyword, MoA_Based_Optimization_Result_File, Compression=True, user_id="", task=None):
    """执行人机协作最终阶段，生成最后的优化草案与表格
    
    参数:
    - Keyword: 主题名称
    - MoA_Based_Optimization_Result_File: MoA 结果文件路径
    - Compression: 是否对论文进行内容压缩
    - user_id: 用户标识
    - task: 任务对象（用于目录归档）
    
    返回:
    - str: 人机协作结果 Markdown 文件路径
    """
    task_id = getattr(task, 'request', {}).get('id', 'default_task_id') if task else 'default_task_id'
    topic_dir = sanitize_folder_name(Keyword.replace(" ", "_"))
    file_path_prefix = fr"{OUTPUT_PATH}/{user_id}/{task_id}/{topic_dir}/MOA"
    file_path_result_prefix = fr"{OUTPUT_PATH}/{user_id}/{task_id}/{topic_dir}/Idea"
    os.makedirs(file_path_prefix, exist_ok=True)
    os.makedirs(file_path_result_prefix, exist_ok=True)
    idea_draft, target_paper_title = extract_message(file=MoA_Based_Optimization_Result_File, split_section="Paper Title")
    moa_table(topic=topic_dir, draft=idea_draft.split("### Summary of the Differences in This Iteration:")[0].strip(), user_id=user_id, task=task)
    _, next_optimization = extract_message_review_moa(file=fr"{file_path_prefix}/{Keyword}_review_moa.md", split_section="Overall Opinions")
    text, optimize_messages = extract_message_review(file=fr"{file_path_prefix}/{Keyword}_review_moa.md", split_section="Iterative Optimization Search Keywords")
    keywords = []
    for optimize_message in optimize_messages.values():
        keywords.append({'keyword': optimize_message})
    target_next_optimization = ""
    for opt in next_optimization:
        target_next_optimization += f"\n{opt}"
    readysearch_key = []
    optimization_keywords = ""
    for keyword in keywords:
        readysearch_key.append(keyword['keyword'])
        optimization_keywords += f"\n{keyword['keyword']}"
    Papers = search_paper(Keywords=readysearch_key, Limit=2)
    title_abstract_prompt = ""
    relatedPaper = []
    for paper in Papers:
        if Compression:
            try:
                compression_result = paper_compression(doi=paper["doi"], title=paper["title"], topic=Keyword, user_id=user_id, task=task)
            except Exception:
                compression_result = "None"
        else:
            compression_result = "None"
        try:
            relatedPaper.append({"title": paper["title"], "abstract": paper["abstract"], "compression_result": compression_result})
        except:
            pass
    if Compression:
        for index, paper in enumerate(relatedPaper):
            title_abstract_prompt += f"\n# The {index + 1} related paper\n## title\n{paper['title']}\n## abstract\n{paper['abstract']}\n## content\n{paper['compression_result']}\n"
    else:
        for index, paper in enumerate(relatedPaper):
            title_abstract_prompt += f"\n# The {index + 1} related paper\n## title\n{paper['title']}\n## abstract\n{paper['abstract']}\n"
    user_prompt = human_ai_collaboration_prompt(target_next_optimization=target_next_optimization, optimization_keywords=optimization_keywords, title_abstract_prompt=title_abstract_prompt, draft=idea_draft)
    with open(fr"{file_path_result_prefix}/{Keyword}_human_ai_collaboration_input.md", 'w', encoding='utf-8') as f:
        f.write(user_prompt)
    idea_iteration_result = call_with_qwenmax(system_prompt="You are a research assistant.", question=user_prompt)
    with open(fr"{file_path_result_prefix}/{Keyword}_human_ai_collaboration_result.md", 'w', encoding='utf-8') as f:
        f.write(idea_iteration_result)
    return fr"{file_path_result_prefix}/{Keyword}_human_ai_collaboration_result.md"
